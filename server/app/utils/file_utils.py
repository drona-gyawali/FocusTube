import csv
import io
import re
from typing import List

import PyPDF2
from openpyxl import load_workbook


def _extract_links_from_text(text: str) -> List[str]:
    """Extracts all URLs from plain text."""
    url_pattern = r"https?://[^\s]+"
    return re.findall(url_pattern, text)


def extract_file_link(file_bytes: bytes, filename: str) -> List[str]:
    """
    Detect file type by extension and extract URLs.
    Supports .txt, .csv, .xlsx, .pdf
    """
    filename = filename.lower()

    if filename.endswith(".txt"):
        text = file_bytes.decode("utf-8", errors="ignore")
        return _extract_links_from_text(text)

    elif filename.endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        links = []
        for row in reader:
            for cell in row:
                links.extend(_extract_links_from_text(cell))
        return links

    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
        links = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        links.extend(_extract_links_from_text(str(cell)))
        return links

    elif filename.endswith(".pdf"):
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        links = []
        for page in pdf_reader.pages:
            text = page.extract_text() or ""
            links.extend(_extract_links_from_text(text))
        return links

    else:
        raise ValueError("Unsupported file type. Supported: .txt, .csv, .xlsx, .pdf")


def extract_file_id(file_url: str) -> str:
    """
    Extract the specific file-id from the long url
    """
    match = re.search(r"/files/([^/]+)/view", file_url)
    if match:
        return match.group(1)
    else:
        return None


def extract_youtube_link_id(url: str) -> str | None:
    """
    Extract the YouTube Video ID from a given URL.
    Returns the video ID string if valid, else None.
    """
    if not url:
        return None

    pattern = (
        r"(?:https?://)?"
        r"(?:www\.)?"
        r"(?:youtube\.com/watch\?v=|"
        r"youtu\.be/|"
        r"youtube\.com/embed/|"
        r"youtube\.com/shorts/)"
        r"([^\s&?/]{11})"
    )

    match = re.search(pattern, url)
    if not match:
        return None

    video_id = match.group(1)

    if re.match(r"^[A-Za-z0-9_-]{11}$", video_id):
        return video_id

    return None
